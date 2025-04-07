#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import datetime
import openpyxl

def analyze_xlsx(file_path):
    """
    分析指定的xlsx文件，将内容导出为txt文件
    
    Args:
        file_path: xlsx文件路径
    
    Returns:
        生成的txt文件路径
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"错误：文件 {file_path} 不存在")
            return None
        
        # 检查文件扩展名
        if not file_path.lower().endswith('.xlsx'):
            print(f"错误：文件 {file_path} 不是xlsx格式")
            return None
        
        # 读取Excel文件
        print(f"正在读取文件: {file_path}")
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # 生成时间戳
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 构建输出文件名
        file_name = os.path.basename(file_path)
        file_name_without_ext = os.path.splitext(file_name)[0]
        output_file = os.path.join(os.path.dirname(file_path), f"{file_name_without_ext}_{timestamp}.txt")
        
        # 将数据写入txt文件
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"Excel文件分析结果: {file_name}\n")
            f.write(f"分析时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*80 + "\n\n")
            
            # 遍历所有工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                f.write(f"工作表: {sheet_name}\n")
                f.write("-"*80 + "\n\n")
                
                # 获取最大行和列
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # 获取标题行（第1行）作为字段名
                headers = []
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=1, column=col)
                    header = str(cell.value) if cell.value is not None else f"列{col}"
                    headers.append(header)
                
                # 从第2行开始遍历输出数据
                for row in range(2, max_row + 1):
                    f.write(f"行 {row-1}:\n")
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell_value = str(cell.value) if cell.value is not None else ""
                        f.write(f"{headers[col-1]}: {cell_value}\n")
                    f.write("-"*80 + "\n\n")
                
                f.write("="*80 + "\n\n")
        
        print(f"分析完成。结果已保存到: {output_file}")
        return output_file
    
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        return None

def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("使用方法: python xlsx_analyzer.py <xlsx文件路径>")
        return
    
    file_path = sys.argv[1]
    analyze_xlsx(file_path)

if __name__ == "__main__":
    main() 